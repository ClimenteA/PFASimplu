from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def make_excel_pretty(excel_path: str):

    wb = load_workbook(excel_path)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter 
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    header_fill = PatternFill(start_color="00895a", end_color="00895a", fill_type="solid")
    
    for cell in ws[1]:
        cell.fill = header_fill

    wb.save(excel_path)
